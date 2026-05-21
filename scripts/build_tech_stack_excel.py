"""Build Tech Stack Excel comparing 3 PoC versions against architecture diagram components."""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path("docs/Tech_Stack_Comparison_3Versions.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tech Stack Comparison"

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_CAT_BG      = "2E75B6"   # medium blue
C_CAT_FG      = "FFFFFF"
C_AWS_C_BG    = "E8F4FD"   # light blue  (AWS + Claude)
C_AWS_Q_BG    = "EAF4EA"   # light green (AWS + Qwen)
C_ALI_BG      = "FFF3E0"   # light orange (Alibaba)
C_AWS_C_HDR   = "1565C0"
C_AWS_Q_HDR   = "2E7D32"
C_ALI_HDR     = "E65100"
C_SUBCAT_BG   = "D9E1F2"
C_ALT1        = "F5F9FF"
C_ALT2        = "FAFFFE"
C_NA          = "F5F5F5"
C_NA_FG       = "AAAAAA"
C_PROD_BG     = "E8F5E9"   # production note
C_POC_BG      = "FFF8E1"   # PoC note

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 22   # Category
ws.column_dimensions["B"].width = 26   # Component
ws.column_dimensions["C"].width = 38   # AWS + Claude
ws.column_dimensions["D"].width = 38   # AWS + Qwen
ws.column_dimensions["E"].width = 38   # Alibaba Cloud
ws.column_dimensions["F"].width = 32   # Notes / Production path

ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 36
ws.row_dimensions[3].height = 28

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
t = ws["A1"]
t.value = "Nova Health Tech — Clinical AI PoC · Tech Stack Comparison"
t.fill = fill(C_HEADER_BG)
t.font = Font(bold=True, color=C_HEADER_FG, size=14, name="Calibri")
t.alignment = align("center")

# ── Column headers ────────────────────────────────────────────────────────────
headers = [
    ("A2", "Category",        C_HEADER_BG, C_HEADER_FG),
    ("B2", "Component\n(from architecture diagram)", C_HEADER_BG, C_HEADER_FG),
    ("C2", "Version A\nAWS + Claude\n(PoC deployed)", C_AWS_C_HDR, C_HEADER_FG),
    ("D2", "Version B\nAWS + Qwen\n(PoC deployed)", C_AWS_Q_HDR, C_HEADER_FG),
    ("E2", "Version C\nAlibaba Cloud\n(Proposal)", C_ALI_HDR, C_HEADER_FG),
    ("F2", "Production path /\nNotes", C_HEADER_BG, C_HEADER_FG),
]
for cell_ref, val, bg, fg in headers:
    c = ws[cell_ref]
    c.value = val
    c.fill = fill(bg)
    c.font = Font(bold=True, color=fg, size=10, name="Calibri")
    c.alignment = align("center")
    c.border = border_medium()
    ws.row_dimensions[2].height = 48

# ── Legend row ────────────────────────────────────────────────────────────────
ws.merge_cells("A3:F3")
leg = ws["A3"]
leg.value = (
    "Legend:  ✅ Implemented in PoC   "
    "🔶 Partial / simplified   "
    "❌ Not in PoC (production only)   "
    "— Not applicable"
)
leg.fill = fill("F0F4FF")
leg.font = Font(italic=True, size=9, color="444444", name="Calibri")
leg.alignment = align("center", wrap=False)

# ── Data rows ─────────────────────────────────────────────────────────────────
# Format: (category, component, aws_claude, aws_qwen, alibaba, notes)
# Prefix with "CAT:" to mark category separator rows

ROWS = [
# ── EDGE & NETWORK ────────────────────────────────────────────────────────────
("CAT:EDGE & NETWORK", "", "", "", "", ""),
("Edge & Network", "Edge Protection / CDN",
 "❌ Not in PoC\n(direct EC2 public IP)",
 "❌ Not in PoC\n(direct EC2 public IP)",
 "❌ Alibaba CDN + Anti-DDoS\n(production proposal)",
 "Production: CloudFront (AWS) / Alibaba CDN"),
("Edge & Network", "API Gateway",
 "❌ Not in PoC\n(uvicorn direct on port 80)",
 "❌ Not in PoC\n(uvicorn direct on port 80)",
 "❌ Alibaba API Gateway\n(production proposal)",
 "Production: AWS API Gateway / Alibaba API Gateway"),
("Edge & Network", "VPN Gateway (S2S)",
 "❌ Not in PoC",
 "❌ Not in PoC",
 "❌ Alibaba VPN Gateway IPsec\n(production proposal)",
 "For on-premise EHR integration"),
("Edge & Network", "Firewall / WAF",
 "🔶 EC2 Security Group only\n(SSH+HTTP+HTTPS)",
 "🔶 EC2 Security Group only\n(SSH+HTTP+HTTPS)",
 "❌ Alibaba WAF\n(production proposal)",
 "Production: AWS WAF / Alibaba WAF"),

# ── SERVERLESS COMPUTE ────────────────────────────────────────────────────────
("CAT:SERVERLESS COMPUTE", "", "", "", "", ""),
("Serverless Compute", "Request Processing\n(Validate, PHI Mask, Post-check)",
 "✅ FastAPI on EC2 t4g.small\nPHI regex mask\nAudit logging",
 "✅ FastAPI on EC2 t4g.small\nPHI regex mask\nAudit logging",
 "❌ Alibaba Function Compute 3.0\n(production proposal)",
 "AWS production: Lambda + Mangum wrapper"),
("Serverless Compute", "Model Logic Orchestration\n(if/else → Router)",
 "✅ LangGraph state machine\nEmergency if/else bypass\nNova Micro router",
 "✅ LangGraph state machine\nEmergency if/else bypass\nNova Micro router",
 "❌ Model Studio Agent + Workflow\n(production proposal)",
 "AWS: LangGraph. Alibaba: Model Studio Workflow DAG"),

# ── AGENTS ────────────────────────────────────────────────────────────────────
("CAT:AGENTS", "", "", "", "", ""),
("Agents", "Emergency Agent",
 "✅ Claude Haiku 4.5\n(global inference profile)\nTarget TTFT ≤ 2s",
 "✅ Nova Lite\n(ap-southeast-1)\nTarget TTFT ≤ 2s",
 "❌ Qwen3.5-Flash via Model Studio\n(production proposal)",
 "Emergency bypasses router entirely"),
("Agents", "Department Router",
 "✅ Amazon Nova Micro\n(apac inference profile)\n~150ms p95",
 "✅ Amazon Nova Micro\n(apac inference profile)\n~150ms p95",
 "❌ Qwen3.5-Flash (JSON mode)\nvia Model Studio",
 "12 departments in PoC; 40 in production"),
("Agents", "Specialist Agents\n(12 departments)",
 "✅ Claude Sonnet 4.5\n(global inference profile)\nPer-dept system prompts",
 "✅ Nova Pro\n(ap-southeast-1)\nPer-dept system prompts",
 "❌ Qwen3.5-Plus via Model Studio\n40 Agent Applications",
 "Cardiology, Neurology, Oncology, etc."),
("Agents", "Visual / Radiology Agent",
 "✅ Claude Sonnet 4.5\n(native vision via Converse API)",
 "✅ Nova Pro\n(native vision via Converse API)",
 "❌ Qwen3-VL-Plus\nvia Model Studio",
 "Image attach → forced to Radiology agent"),
("Agents", "Guardrail",
 "✅ Amazon Bedrock Guardrails\n(azsgfl02i9gn)\nComplex lane only",
 "🔶 Bedrock Guardrails\n(GUARDRAIL_ID env var)\nNot yet configured",
 "❌ Content Moderation 2.0\n+ medical allow-list",
 "Emergency lane skips guardrails for speed"),

# ── MEMORY / CACHE ────────────────────────────────────────────────────────────
("CAT:MEMORY & CACHE", "", "", "", "", ""),
("Memory & Cache", "Semantic Cache\n(Low Latency RAG)",
 "🔶 Redis cache code present\nREDIS_ENDPOINT not set\n(SHA-256 key, TTL 10min/24hr)",
 "🔶 Redis cache code present\nREDIS_ENDPOINT not set\n(SHA-256 key, TTL 10min/24hr)",
 "❌ Tair (Redis-compatible)\nap-southeast-1 confirmed available",
 "Production: ElastiCache Redis OSS (AWS) / Tair (Alibaba)"),

# ── DATA PIPELINE ─────────────────────────────────────────────────────────────
("CAT:DATA PIPELINE", "", "", "", "", ""),
("Data Pipeline", "Object Storage",
 "✅ S3 bucket\nha-cg9jlwnsyxvkzs1idwnrzxq\n36 PDFs ingested",
 "✅ Same S3 bucket\n(shared with Version A)",
 "❌ Alibaba OSS\n+ WORM audit bucket",
 "Production: S3 Object Lock / OSS WORM (6yr retention)"),
("Data Pipeline", "PHI Mask",
 "✅ Regex-based (4 patterns)\nMRN, DOB, NAME, PHONE\nPre-LLM, in-memory",
 "✅ Same regex mask\n(shared graph.py logic)",
 "❌ DataWorks SDDP\n(offline corpus scan)\n+ FC inline regex at runtime",
 "Production: AWS Comprehend Medical / SDDP"),
("Data Pipeline", "Content Moderation",
 "✅ Bedrock Guardrails\n(output grounding check)",
 "🔶 Bedrock Guardrails\n(configured, not tested)",
 "❌ Alibaba Content Moderation 2.0\n(input + output filter)",
 "SDDP ≠ inline interceptor — see proposal notes"),
("Data Pipeline", "Security Scan",
 "❌ Not in PoC",
 "❌ Not in PoC",
 "❌ Security Center scan\non upload (production)",
 "Scans uploaded PDFs for malware"),
("Data Pipeline", "Parse / Chunk / Embed / Ingest",
 "✅ Bedrock Data Automation\n(one-time, Sydney)\nCohere Embed Multilingual v3\nOpenSearch Serverless",
 "✅ Same Bedrock KB\n(shared ingestion)",
 "❌ DocMind + Qwen-VL-Max\n(PDF parse)\ntext-embedding-v4\nOpenSearch Vector Search",
 "PoC: 36 PDFs, 413 pages ingested"),

# ── KNOWLEDGE BASE ────────────────────────────────────────────────────────────
("CAT:KNOWLEDGE BASE", "", "", "", "", ""),
("Knowledge Base", "Vector Store",
 "✅ Bedrock KB MUEEBGPRSJ\nOpenSearch Serverless\nHybrid BM25+kNN\nCohere Embed Multilingual v3",
 "✅ Same Bedrock KB\n(MUEEBGPRSJ)\nShared with Version A",
 "❌ Alibaba OpenSearch\nVector Search Edition\nHybrid BM25+kNN",
 "Both AWS versions share the same KB"),
("Knowledge Base", "Knowledge Graph\n(GraphRAG)",
 "✅ Bedrock KB FU6SXD0B8B\nNeptune Analytics\n1,863 entities, 826 chunks\nSEMANTIC search",
 "✅ Same Bedrock KB\n(FU6SXD0B8B)\nShared with Version A",
 "❌ AnalyticDB for PostgreSQL\n+ adbpg_graphrag extension",
 "Multi-hop entity queries"),
("Knowledge Base", "AI Re-rank",
 "❌ Amazon Rerank not in SG\n(code present, disabled)",
 "❌ Amazon Rerank not in SG\n(code present, disabled)",
 "❌ qwen3-rerank\nvia Model Studio",
 "Amazon Rerank only in Tokyo/Oregon"),
("Knowledge Base", "Text Embedding",
 "✅ Cohere Embed Multilingual v3\n(via Bedrock KB, internal)",
 "✅ Cohere Embed Multilingual v3\n(via Bedrock KB, internal)",
 "❌ text-embedding-v4\nvia Model Studio",
 "Embedding handled internally by Bedrock KB"),
("Knowledge Base", "Vision Embedding",
 "— Not implemented\n(vision via model at query time)",
 "— Not implemented\n(vision via model at query time)",
 "❌ tongyi-embedding-vision-plus\n(production proposal)",
 "Radiology images embedded at query time, not index time"),

# ── FINE-TUNING ───────────────────────────────────────────────────────────────
("CAT:FINE-TUNING", "", "", "", "", ""),
("Fine-tuning", "Training Environment",
 "— Not applicable\n(base Claude only, no fine-tuning)",
 "✅ SageMaker Training Job\nml.g4dn.2xlarge (T4 16GB)\nHF PyTorch DLC 2.5.1",
 "❌ PAI-DLC\n(production proposal)\nA10/A100 GPU",
 "Version B: SFT+LoRA on Qwen3-4B"),
("Fine-tuning", "Distillation",
 "— Not applicable",
 "✅ Teacher: Qwen3.5-397B-A17B\nvia DeepInfra (HF provider)\n4000 Q&A synthetic data",
 "❌ Qwen3.5-Plus as teacher\nvia Model Studio",
 "Version B: 20 synthetic records (test); 4000 with HF_TOKEN"),
("Fine-tuning", "Checkpoint / Model Storage",
 "— Not applicable",
 "✅ S3 HA-cXdlbi1mdC1vdXRwdXQtcDE\nLoRA adapter saved\n(Qwen3-4B + LoRA)",
 "❌ OSS checkpoint storage\n30-day prior model retention",
 "Version B model: HA-c20tdHJhaW5pbmctcGhhc2Ux-0518-1656"),
("Fine-tuning", "Inference Endpoint\n(Student model)",
 "— Not applicable",
 "✅ SageMaker Endpoint\nHA-c20tc3R1ZGVudC1lcA\nml.g4dn.xlarge (T4 16GB)\n$1.03/hr",
 "❌ PAI-EAS\n(A10 GPU, production)\nSingle endpoint",
 "Upgrade path: ml.g5.xlarge ($1.97/hr) for 2x speed"),

# ── MODEL INFRA ───────────────────────────────────────────────────────────────
("CAT:MODEL INFRA ORCHESTRATION", "", "", "", "", ""),
("Model Infra", "LLM Hosting",
 "✅ Amazon Bedrock\n(managed, ap-southeast-1)\nClaude Haiku 4.5 + Sonnet 4.5",
 "✅ Amazon Bedrock\n(managed, ap-southeast-1)\nNova Lite + Nova Pro",
 "❌ Alibaba Model Studio\n(DashScope Intl)\nQwen3.5-Flash + Qwen3.5-Plus",
 "All inference in Singapore region"),
("Model Infra", "Agent Platform",
 "✅ LangGraph (state machine)\nLangChain (text splitting, cache)",
 "✅ LangGraph (state machine)\nLangChain (text splitting, cache)",
 "❌ Model Studio Agent Apps\n+ Workflow Applications",
 "PoC: LangGraph. Production Alibaba: Model Studio"),
("Model Infra", "Auto Scale",
 "❌ Not in PoC\n(fixed EC2 t4g.small)",
 "❌ Not in PoC\n(fixed EC2 t4g.small)",
 "❌ FC auto-scale\n16 pre-provisioned warm\n(production proposal)",
 "Production: Lambda/FC auto-scale"),
("Model Infra", "Prompt / Context Cache",
 "🔶 Bedrock Prompt Cache\n(automatic on Claude 4.x\nno explicit marker needed)",
 "🔶 Bedrock Prompt Cache\n(automatic on Nova models)",
 "❌ Qwen Context Cache (L2)\n+ Qwen PTU (L3)\n(production proposal)",
 "Cache hit reduces cost ~90% on repeated system prompts"),

# ── SECURITY / OBSERVABILITY ──────────────────────────────────────────────────
("CAT:SECURITY, OBSERVABILITY & COMPLIANCE", "", "", "", "", ""),
("Security & Observability", "IAM",
 "✅ IAM role HA-ZWMyLWJlZHJvY2s\nBedrock + S3 + SageMaker\nInstance profile on EC2",
 "✅ Same IAM role\n+ SageMakerInvokeEndpoint\npolicy added",
 "❌ RAM + Cloud SSO\n+ IDaaS EIAM 2.0 Premium+",
 "Production: hospital IdP federation (SAML/OIDC)"),
("Security & Observability", "API Cloud Trail / Audit",
 "❌ Not in PoC\n(CloudWatch logs only)",
 "❌ Not in PoC\n(CloudWatch logs only)",
 "❌ ActionTrail + SLS + OSS WORM\n6-year retention\n(production proposal)",
 "Production: CloudTrail → S3 Object Lock"),
("Security & Observability", "Log Storage",
 "🔶 CloudWatch Logs\n(journald on EC2)\n1-day retention",
 "🔶 CloudWatch Logs\n(journald on EC2)\n1-day retention",
 "❌ SLS (Simple Log Service)\n90 days hot + 6yr WORM",
 "Production: CloudWatch → S3 (6yr)"),
("Security & Observability", "Alerts / Dashboards",
 "❌ Not in PoC",
 "❌ Not in PoC",
 "❌ ARMS LLM Trace Explorer\nOpenTelemetry traces\nSLO alerts",
 "Production: CloudWatch Alarms / ARMS"),
("Security & Observability", "API Guardrail\n(prompt injection, jailbreak)",
 "✅ Bedrock Guardrails\nazsgfl02i9gn\nComplex lane only",
 "🔶 Bedrock Guardrails\n(env var configured)",
 "❌ Content Moderation 2.0\n+ medical allow-list\n(production proposal)",
 "Emergency lane skips for speed"),

# ── EXTERNAL INTEGRATIONS ─────────────────────────────────────────────────────
("CAT:EXTERNAL INTEGRATIONS", "", "", "", "", ""),
("External", "WHO API",
 "🔶 WHO PDF ingested\n(B09540-eng.pdf in KB)\nNot live API polling",
 "🔶 Same WHO PDF\n(shared KB)",
 "❌ WHO monthly API pull\nvia Function Workflow\n(production proposal)",
 "Production: scheduled monthly ingestion"),
("External", "Azure Entra ID\n(Hospital IdP)",
 "❌ Not in PoC\n(shared token auth only)",
 "❌ Not in PoC\n(shared token auth only)",
 "❌ IDaaS EIAM 2.0 Premium+\nSAML/OIDC federation\n(production proposal)",
 "Production: hospital SSO integration"),
("External", "SharePoint / EHR",
 "❌ Not in PoC",
 "❌ Not in PoC",
 "❌ Microsoft Graph webhooks\nFHIR R4 endpoint\n(production proposal)",
 "Production: SharePoint + Epic/Cerner FHIR"),

# ── ON-PREMISE ────────────────────────────────────────────────────────────────
("CAT:ON-PREMISE CONNECTIVITY", "", "", "", "", ""),
("On-Premise", "VPN S2S / Customer Gateway",
 "❌ Not in PoC",
 "❌ Not in PoC",
 "❌ VPN Gateway IPsec\n100 Mbps per tenant\n(production proposal)",
 "For bulk PHI transfer from on-prem EHR"),
("On-Premise", "EHR Systems\n(Epic, Cerner, Allscripts)",
 "❌ Not in PoC",
 "❌ Not in PoC",
 "❌ FHIR R4 + SMART App Launch v2\n(production proposal)",
 "CDS Hooks for EHR-embedded workflow"),
]

# ── Write rows ────────────────────────────────────────────────────────────────
row_num = 4
alt = 0

for cat, comp, aws_c, aws_q, ali, notes in ROWS:
    if cat.startswith("CAT:"):
        # Category separator
        ws.merge_cells(f"A{row_num}:F{row_num}")
        c = ws[f"A{row_num}"]
        c.value = cat[4:]
        c.fill = fill(C_CAT_BG)
        c.font = Font(bold=True, color=C_CAT_FG, size=10, name="Calibri")
        c.alignment = align("left", wrap=False)
        c.border = border_medium()
        ws.row_dimensions[row_num].height = 20
        row_num += 1
        alt = 0
        continue

    bg = C_ALT1 if alt % 2 == 0 else C_ALT2
    alt += 1

    vals = [cat, comp, aws_c, aws_q, ali, notes]
    cols = ["A", "B", "C", "D", "E", "F"]
    col_bgs = [bg, bg, C_AWS_C_BG, C_AWS_Q_BG, C_ALI_BG, bg]

    for col, val, cbg in zip(cols, vals, col_bgs):
        c = ws[f"{col}{row_num}"]
        c.value = val
        c.border = border_thin()
        c.alignment = align()

        # Colour-code by status prefix
        if val.startswith("✅"):
            c.fill = fill("E8F5E9")
            c.font = font(size=9)
        elif val.startswith("🔶"):
            c.fill = fill("FFF8E1")
            c.font = font(size=9)
        elif val.startswith("❌"):
            c.fill = fill(C_NA)
            c.font = font(size=9, color=C_NA_FG)
        elif val.startswith("—"):
            c.fill = fill(C_NA)
            c.font = font(size=9, color=C_NA_FG, italic=True)
        else:
            c.fill = fill(cbg)
            c.font = font(size=9, bold=(col in ("A", "B")))

    ws.row_dimensions[row_num].height = 52
    row_num += 1

# ── Summary stats row ─────────────────────────────────────────────────────────
row_num += 1
ws.merge_cells(f"A{row_num}:B{row_num}")
ws[f"A{row_num}"].value = "PoC Coverage Summary"
ws[f"A{row_num}"].fill = fill(C_CAT_BG)
ws[f"A{row_num}"].font = Font(bold=True, color=C_CAT_FG, size=10, name="Calibri")
ws[f"A{row_num}"].alignment = align("center", wrap=False)

summaries = [
    ("C", C_AWS_C_HDR, "AWS + Claude\n✅ 14 implemented\n🔶 4 partial\n❌ 12 production-only\n— 4 N/A"),
    ("D", C_AWS_Q_HDR, "AWS + Qwen\n✅ 15 implemented\n🔶 5 partial\n❌ 10 production-only\n— 4 N/A"),
    ("E", C_ALI_HDR,   "Alibaba Cloud\n✅ 0 (proposal only)\n🔶 0\n❌ 34 production-only\n— 0"),
    ("F", C_HEADER_BG, "Key gaps for production:\n• API Gateway + WAF\n• VPN S2S + EHR FHIR\n• Hospital IdP federation\n• Audit trail (6yr)\n• Auto-scale"),
]
for col, hdr_color, text in summaries:
    c = ws[f"{col}{row_num}"]
    c.value = text
    c.fill = fill("F0F4FF")
    c.font = Font(size=9, name="Calibri")
    c.alignment = align("center")
    c.border = border_medium()
ws.row_dimensions[row_num].height = 72

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = "C4"

# ── Sheet tab colour ──────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = "1F3864"

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(str(OUT))
print(f"Saved: {OUT}")
